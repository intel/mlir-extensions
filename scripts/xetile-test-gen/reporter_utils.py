
import pandas as pd
from openpyxl import load_workbook

def add_columns_right(worksheet, column_names):
    # Find the first empty cell in the first row (i.e., no column name)
    for i in range(1, worksheet.max_column + 3):
        if worksheet.cell(row=1, column=i).value is None:
            for col_idx,name in enumerate(column_names):
                # print(f"Adding column {name}")
                worksheet.cell(row=1, column=i+col_idx, value=name)
            break

def collect_formulas(excel_path):
        book = load_workbook(excel_path, data_only=False)
        sheet = book.active
        formulas = {}
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == 'f':
                    formulas[cell.coordinate] = cell.value
        return formulas

def reapply_formulas(worksheet, formulas):
    for cell_coordinate, formula in formulas.items():
        worksheet[cell_coordinate].value = formula

def match_dtypes_to_first(df1, df2, cols):
    for col in cols:
        df2[col] = df2[col].astype(df1[col].dtypes.name)
    return df2
