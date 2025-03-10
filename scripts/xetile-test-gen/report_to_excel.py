import re
import glob
import os
import argparse
import traceback
import pandas as pd
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

import reporter_utils

DEFAULT_TEST_PARAMS = "BatchSize,M,K,N,wgm,wgn,sgm,sgn,sgk".split(',')
DEFAULT_SPREADSHEET_NAME = "speedup-report.xlsx"

class XLSXReporter:
    def __init__(self, reporter_args, test_cases_param_list):
        self.report_path = reporter_args.report_name
        self.sheet_name = reporter_args.sheet_name
        self.reports_dir = reporter_args.reports_dir
        self.test_cases_param_list = test_cases_param_list

    def parse_report_to_dataframe(self, file_path, code_variant, test_delimeter="Testing"):
        with open(file_path, 'r') as f:
            content = f.read()
        tests = content.split(test_delimeter)
        csv_header = self.test_cases_param_list + [f"Best {code_variant} time(ms)"]#, f"{code_variant} ALLCLOSE"]
        data = []
        for test in tests:
            # print(test)
            # test_case_line = re.search(r'Kernel test_kernel : (.*) registers', test)
            # if not test_case_line:
            test_name = re.search(r'sg_gemm_(\d+)_(\d+)x(\d+)x(\d+)_wgm(\d+)_wgn(\d+)_sgm(\d+)_sgn(\d+)_sgk(\d+)', test)
            if not test_name:
                continue
            test_case = ""
            for i in range(1, 10):
                test_case += f"{test_name.group(i)},"
            test_case = test_case[:-1]
            # else:
            #     test_case_params = test[test_case_line.span()[1]+1:]
            #     test_case_params = re.search(r'( .* )', test_case_params)
            #     test_case = test_case_params.group(1).strip()
            kernel_time_line = re.search(r':avg: .*, min: (.*), max: .*', test)
            min_time = None
            # allclose = None
            if kernel_time_line:
                # allclose_line = re.search(r'\[ALLCLOSE: (.*)\]', test)
                # allclose = allclose_line.group(1).strip().lower()
                min_time = kernel_time_line.group(1).strip()

            data.append(test_case.split(',') + [min_time])#, allclose])
        result_df = pd.DataFrame(data, columns=csv_header)
        # print(result_df)

        for col in self.test_cases_param_list:
            result_df[col] = result_df[col].astype('uint64')
        result_df[csv_header[-1]] = result_df[csv_header[-1]].astype('float64')

        return result_df

    def create_excel_spreadsheet_layout_new(self):
        wb = Workbook()
        ws = wb.active
        col_names_letters = {self.test_cases_param_list[i-1]: get_column_letter(i) for i in range(1, len(self.test_cases_param_list) + 1) }
        for col in self.test_cases_param_list:
            ws[f"{col_names_letters[col]}1"] = f"{col}"
        wb.save(self.sheet_name)
        return wb

    def generate_tflops_speedup(self, worksheet, time_colname, tflops_speedup_colnames, speedup=False, nrows=300):
        colname_letter_map = {worksheet.cell(row=1, column=i).value: get_column_letter(i) for i in range(1, worksheet.max_column+1)}
        tflops_col_letter = colname_letter_map[tflops_speedup_colnames[0]]
        for row_idx in range(2, nrows + 2):
            # TFLOPS: BatchSize * (2 * B2 * C2 * D2) / (K2 * 10^12)
            formula_TFLOPS_baseline = f"""= {colname_letter_map["BatchSize"]}{row_idx} * (2 * {colname_letter_map["M"]}{row_idx} * {colname_letter_map["K"]}{row_idx} * {colname_letter_map["N"]}{row_idx})"""
            formula_TFLOPS_baseline += " / "
            formula_TFLOPS_baseline += f"({colname_letter_map[time_colname]}{row_idx} / 1000 * 10^12)"
            worksheet[f'{tflops_col_letter}{row_idx}'] = f"{formula_TFLOPS_baseline}"
            if speedup:
                formula_speedup = f"""= ROUND({tflops_col_letter}{row_idx} / {colname_letter_map["baseline TFLOPS"]}{row_idx}, 2)"""
                worksheet[f'{colname_letter_map[tflops_speedup_colnames[1]]}{row_idx}'] = f"{formula_speedup}"
        if speedup:
            for speedup_col in colname_letter_map.keys():
                if speedup_col.endswith("speedup"):
                    speedup_col_letter = colname_letter_map[speedup_col]
                    worksheet.conditional_formatting.add(f"{speedup_col_letter}{1}:{speedup_col_letter}{nrows}",
                        ColorScaleRule(start_type='num', start_value=0, start_color='FF6666',
                        mid_type='num', mid_value=1, mid_color='FFFF66',
                        end_type='num', end_value=4, end_color='66CC00')
                    )
                    col = worksheet.column_dimensions[f"{speedup_col_letter}"]
                    col.font = Font(bold=True)

    def update_excel_with(self, measurements_df, code_variant, speedup=False):
        spreadsheet_df = pd.read_excel(self.sheet_name)
        reporter_utils.match_dtypes_to_first(spreadsheet_df, measurements_df, self.test_cases_param_list)
        if not spreadsheet_df.empty:
            formulas = reporter_utils.collect_formulas(self.sheet_name)
        writer = pd.ExcelWriter(self.sheet_name, engine='openpyxl', mode='w')
        workbook = writer.book
        merged_df = pd.merge(spreadsheet_df, measurements_df, on=self.test_cases_param_list, how='outer', suffixes=('old', ''))
        conflicting_columns = [col for col in merged_df.columns if col.endswith('old') or col.startswith(code_variant)]
        merged_df.drop(columns=conflicting_columns, inplace=True)
        merged_df.to_excel(writer, index=False)
        worksheet = workbook.active
        if not spreadsheet_df.empty:
            reporter_utils.reapply_formulas(worksheet, formulas)
        new_derived_cols = [f"{code_variant} TFLOPS"]
        if speedup:
            new_derived_cols.append(f"{code_variant} speedup")
        reporter_utils.add_columns_right(worksheet, new_derived_cols)
        self.generate_tflops_speedup(worksheet, f"Best {code_variant} time(ms)", new_derived_cols, speedup)
        workbook.save(self.sheet_name)

    def to_excel_table(self):
        # General workflow:
        # 1. Create empty spreadsheet with predetermined structure: test_cases_param_list
        # 2. Fill the spreadsheet with baseline measurements.
        # 3. Update the spreadsheet with code version (prefetch/large_loads/...) measurements:
            # 3.1: Collect all formulas in the spreadsheet
            # 3.2: Load the spreadsheet in Pandas df
            # 3.3: Parse the test reports into Pandas df
            # 3.4: Perform df.merge on test_cases_param_list
            # 3.5: Save merged df as spreadsheet
            # 3.5: Reapply formulas (original columns are not modified or deleted, so it's)
            # 3.6: Add new formula for code version's TFLOPS and speedup compared to Baseline
            # 3.7: Save the spreadsheet

        original_workbook = load_workbook(self.sheet_name, data_only=False) if os.path.exists(self.sheet_name) else None
        try:
            if self.report_path is not None:
                # Update existing table specified in self.report_path: adds column to the right and creates speedup & TFLOPS columns.
                filename, _ = os.path.splitext(os.path.basename(self.report_path))
                parsed_df = self.parse_report_to_dataframe(self.report_path, filename)
                self.update_excel_with(parsed_df, filename, speedup=True)
            elif self.reports_dir is not None:
                # We create a new table from all reports in the self.reports_dir directory
                baseline = None
                reports = []
                for file in glob.glob(f"{self.reports_dir}/*.txt"):
                    if file.endswith("baseline.txt"):
                        baseline = file
                    else:
                        reports.append(file)
                if baseline is None:
                    raise Exception("You must provide baseline report (baseline.txt) to construct speedup")
                # Empty spreadsheet
                self.create_excel_spreadsheet_layout_new()
                # Initialize baseline
                parsed_df = self.parse_report_to_dataframe(baseline, "baseline")
                self.update_excel_with(parsed_df, "baseline")
                # Add all other code variants (e.g., prefetch, large loads, etc.) with relative speedup
                for file in reports:
                    filename, _ = os.path.splitext(os.path.basename(file))
                    parsed_df = self.parse_report_to_dataframe(file, filename)
                    self.update_excel_with(parsed_df, filename, speedup=True)
        except Exception as e:
            print(traceback.format_exc())
            if original_workbook is not None:
                print("Error, saving the original workbook")
                original_workbook.save(self.sheet_name)

if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--reports_dir', '-f', help="Default: current directory. Will create new excel spreadsheet for all files in the \
                         given directory, make sure to have report baseline.txt", type=str, default="./")
    parser.add_argument('--report_name', '-r', help="Filename of the report used to update the spreadsheet", type=str, default=None)
    parser.add_argument('--sheet_name', '-s', help="Name of the spreadsheet that will be created/updated (use in combination \
                         with either --reports_dir or --report_name)", type=str, default=DEFAULT_SPREADSHEET_NAME)
    args = parser.parse_args()
    reporter = XLSXReporter(args, DEFAULT_TEST_PARAMS)
    reporter.to_excel_table()
